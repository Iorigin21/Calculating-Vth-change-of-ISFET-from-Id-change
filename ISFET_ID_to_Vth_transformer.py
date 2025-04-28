"""
Created on Sun June 15 2022
@author: Yingtao Yu
Program for converting ID-time data to Vth-change Vs. time data
Input file: ID-time data(sampling.txt) measured by Hp4155A semiconductor analyzer, ID-VG(IV.txt) data of ISFETs under test,
            time(time.txt) for defining the starting points of the monitoring.
Output file: Vt1.xlsx is the calculated Vth-change Vs. time data
"""

import numpy as np
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import xlrd
from openpyxl import load_workbook

#Filepath to save ID-time data and ID-VG data
FilePath = r"D:\PHD DATA\ISFET-array\Electrical Test\1"

Test = []
IVfile = []
Itfile = []
Begin = []
Vtfile = []
color = ['b','r','c','y','k','g','m','w']

for foldername in os.listdir(FilePath):
    Test.append(foldername)
    fpt = os.path.join(FilePath, foldername)
    for i in os.listdir(fpt):
        if i.split(".")[1] =="txt":
            path = fpt + '/' + i
            r = open(path)
            readl = r.readline()
            if len(readl.split(" ")) > 1 and readl.split(" ")[1] == "IV":
                IVfile.append(path)
            if len(readl.split(" ")) == 1:
                Begin.append(float(readl.split(" ")[0]))
            else:
                Itfile.append(path)

    workbook = Workbook()
    workbook.save(fpt + '/' + 'IV.xlsx')
    workbook.save(fpt + '/' + 'It.xlsx')
    workbook.save(fpt + '/' + 'Vt.xlsx')
    workbook.save(fpt + '/' + 'Vt1.xlsx')

    IVxls = fpt + '/' + 'IV.xlsx'
    Itxls = fpt + '/' + 'It.xlsx'
    Vtxls = fpt + '/' + 'Vt.xlsx'
    Vtxls1 = fpt + '/' + 'Vt1.xlsx'
    Vtfile.append(Vtxls1)

    with pd.ExcelWriter(IVxls) as writer:

        f = open(IVfile[-1])
        j = int(1)  # rownumber
        Vg = []
        Id = []
        Is = []
        Ig = []
        Deviceid = []
        t = 0
        while True:
            lines = f.readline()
            if lines == "":
                break
            if lines.split(" ")[0] == "IV":
                Deviceid.append(lines.split("[")[1].split("]")[0])
                t = j
                while True:
                    lines = f.readline()
                    if t != 0 and j >= t + 4 and lines != "\n" and lines != "":
                        Vg.append(float(lines.split("\t")[0]))
                        Id.append(float(lines.split("\t")[1]))
                        Is.append(float(lines.split("\t")[2]))
                        Ig.append(float(lines.split("\t")[3]))
                        j += 1
                    if t != 0 and j <= t + 3:
                        j += 1
                    if t != 0 and j >= t + 4 and (lines == "\n" or lines == ""):
                        Save = pd.DataFrame({'Vg': Vg, 'Is': Is, 'Id': Id, 'Ig': Ig})
                        Save.to_excel(writer, sheet_name=Deviceid[-1])
                        for i in (Vg, Ig, Is, Id):
                            i.clear()
                        break
            j += 1

    with pd.ExcelWriter(Itxls) as writer:
        f = open(Itfile[-1])
        j = int(1)  # rownumber
        Tt = []
        Id = []
        Is = []
        Ig = []
        Tt1 = []
        Id1 = []
        Is1 = []
        Ig1 = []
        Deviceid1 = []
        t = 0
        tt = 0
        while True:
            lines = f.readline()
            if lines == "":
                break
            if lines.split(" ")[0] == "Time":
                for i in lines.split("["):
                    if list(i)[0] != "T":
                        Deviceid1.append(i.split("]")[0])
                        print(Deviceid1)
                        dn = len(Deviceid1)
            if j >= 15:
                for i in Deviceid1:
                    Tt.append(lines.split("\t")[t * 5])
                    Is.append(lines.split("\t")[t * 5 + 1])
                    Id.append(lines.split("\t")[t * 5 + 2])
                    Ig.append(lines.split("\t")[t * 5 + 4])
                    t += 1
                t = 0
            j += 1
        for i in Deviceid1:
            while True:
                if t >= len(Tt) / dn:
                    Save = pd.DataFrame({'Time': Tt1, 'Is': Is1, 'Id': Id1, 'Ig': Ig1})
                    Save.to_excel(writer, sheet_name=i)
                    for i in (Tt1, Ig1, Is1, Id1):
                        i.clear()
                    t = 0
                    break
                Tt1.append(Tt[dn * t + tt])
                Is1.append(Is[dn * t + tt])
                Id1.append(Id[dn * t + tt])
                Ig1.append(Ig[dn * t + tt])
                t += 1
            tt += 1

    with pd.ExcelWriter(Vtxls) as writer:
        for j in Deviceid1:
            Vgfit = []
            IVdata = pd.read_excel(IVxls, usecols=[1, 3], sheet_name=j)
            Itdata = pd.read_excel(Itxls, usecols=[1, 3], sheet_name=j)
            Ttt = Itdata['Time'].values.tolist()
            CurrT = Itdata['Id'].values.tolist()
            Ttt1 = [float(i) for i in Ttt]
            CurrT1 = [abs(float(i)) for i in CurrT]

            '''plt.plot(Ttt1, CurrT1)
            plt.yscale('log')
            plt.show()'''

            Volt = IVdata['Vg'].values.tolist()
            Curr = IVdata['Id'].values.tolist()
            Volt11 = [float(i) for i in Volt]
            Curr11 = [float(i) for i in Curr]
            Id.clear()
            Vg.clear()
            Vgfit.clear()
            t = 0
            Vt = []
            Vt.clear()
            for i in Volt:
                if float(Curr[ int(t) ]) < min(CurrT1) and float(Curr[t+1]) >= min(CurrT1):
                    tmin = t
                if float(Curr[t]) <= max(CurrT1) and float(Curr[t+1]) > max(CurrT1):
                    tmax =t+1
                t += 1

            for i in Volt[tmin-1:tmax+1]:
                Vg.append(float(i))
            for i in Curr[tmin-1:tmax+1]:
                Id.append(float(i))

#----------------------------- Polyfit bad -------------------------------------------------------------------
            '''deg = np.polyfit(Id, Vg, 3)  # the order of fitting is really important. Not the higher the better
            fit = np.poly1d(deg)

            Vt = fit(CurrT1)'''

#----------------------------- Curve Log fitting Good -------------------------------------------------------
            '''initial guess value, using equation: VG= a*log(ID-c)+b to fit ID-VG data under testing Id range'''
            a_guess = 0.01
            b_guess = 0.1
            c_guess = -1E-10


            popt, pcov = curve_fit(lambda l, a, b, c: a * np.log(l - c) + b, Id, Vg,  p0=(a_guess, b_guess, c_guess), maxfev=50000)
            a = popt[0]
            b = popt[1]
            c = popt[2]
            Vt = a * np.log(CurrT1 - c) + b

            print(a,b,c)

            Save = pd.DataFrame({'Time': Ttt1, 'Vt': Vt})
            Save.to_excel(writer, sheet_name=j)


#-------------------------------Check VG-ID fitting result---------------------------------------
            Idplot = []
            for i in range(0,1000):
                Idplot.append(min(Id)+i*((max(Id)-min(Id))/1000))
            plt.plot(Id, Vg)
            plt.plot(Idplot, a * np.log(Idplot - c) + b, color='r', label='fit', linestyle='dashed')
            #plt.plot(Curr11, Volt11)
            #plt.plot(Curr11, fit(Curr11))
            plt.axvline(x=min(CurrT1))
            plt.axvline(x=max(CurrT1))
            plt.ylim(ymax=max(Vg)+0.2*(max(Vg)-min(Vg)))
            plt.ylim(ymin=min(Vg)-0.2*(max(Vg)-min(Vg)))
            plt.xscale('log')
            plt.show()
            Idplot.clear()
#-----------------------------------Normalizaion---------------------------------------#
    with pd.ExcelWriter(Vtxls1) as writer:
        List_t = []
        List_v = []

        for j in Deviceid1:
            Vtdata = pd.read_excel(Vtxls, usecols=[1, 2], sheet_name=j)
            Vtt = Vtdata['Time'].values.tolist()
            Vtv = Vtdata['Vt'].values.tolist()
            for i in Vtt:
                if i >= Begin[-1] and Vtt[Vtt.index(i) - 1] < Begin[-1]:
                    print(i)
                    Vtt1 = (np.array(Vtt) - i).tolist()
                    Vtv1 = (np.array(Vtv) - Vtv[Vtt.index(i)]).tolist()
            List_t.append(Vtt1)
            List_v.append(Vtv1)
            Vtt2 = [i / 60 for i in Vtt1]
            Vtv2 = [i * 1000 for i in Vtv1]
            Save = pd.DataFrame({'T-T0': Vtt2, 'Vt-Vt0': Vtv2})
            Save.to_excel(writer, sheet_name=(j))
            Vtt.clear()
            Vtv.clear()
            Vtt2.clear()
            Vtv2.clear()

#-----------------------------------display data plot--------------------------------------

for i in Vtfile:
    wb = load_workbook(i, read_only=True, keep_links=False)
    for j in wb.sheetnames:
        Vtdata = pd.read_excel(i, usecols=[1, 2], sheet_name=j)
        Vtt = Vtdata['T-T0'].values.tolist()
        Vtv = Vtdata['Vt-Vt0'].values.tolist()
        #Vtt1 = [i/60 for i in Vtt]
        #Vtv1 = [i*1000 for i in Vtv]
        plt.plot(Vtt,Vtv, color=color[Vtfile.index(i)], label = (i.split("/")[-2] + j))


plt.legend()
plt.xlim(xmax=90)
plt.xlim(xmin=-10)
plt.ylim(ymax=150)
plt.ylim(ymin=-10)
plt.show()