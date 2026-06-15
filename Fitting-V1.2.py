import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from openpyxl import Workbook, load_workbook

FilePath = r"C:\PHD\PHD DATA\1"

Vtfile = []
color = ['b', 'r', 'c', 'y', 'k', 'g', 'm', 'orange']


def classify_txt_files(folder_path):
    """
    在每个子文件夹中识别：
    1. IV sweep 文件
    2. Time/sampling 文件
    3. Begin 时间文件（只含一个数字）
    """
    iv_path = None
    it_path = None
    begin_value = None

    for fname in os.listdir(folder_path):
        path = os.path.join(folder_path, fname)

        if not os.path.isfile(path):
            continue
        if not fname.lower().endswith(".txt"):
            continue

        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            first_line = f.readline().strip()

        parts = first_line.split()

        # 只含一个数字，视为 Begin 文件
        if len(parts) == 1:
            try:
                begin_value = float(parts[0])
                continue
            except ValueError:
                pass

        # IV 文件判定
        if len(parts) > 1 and parts[1] == "IV":
            iv_path = path
        else:
            # 其余 txt 视为时间采样文件
            it_path = path

    if iv_path is None:
        raise FileNotFoundError(f"[{folder_path}] 没找到 IV 文件")
    if it_path is None:
        raise FileNotFoundError(f"[{folder_path}] 没找到 Time/sampling 文件")
    if begin_value is None:
        raise FileNotFoundError(f"[{folder_path}] 没找到 Begin 时间文件")

    return iv_path, it_path, begin_value


def parse_iv_to_excel(iv_path, out_excel):
    """
    解析 IV.txt -> IV.xlsx
    每个 device 一个 sheet
    """
    device_ids = []

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        with open(iv_path, "r", encoding="utf-8", errors="ignore") as f:
            Vg, Id, Is, Ig = [], [], [], []

            while True:
                line = f.readline()
                if line == "":
                    break

                parts = line.strip().split()

                if len(parts) >= 4 and parts[0] == "IV":
                    # 保存上一个 device（如果有）
                    if len(Vg) > 0 and len(device_ids) > 0:
                        df = pd.DataFrame({
                            'Vg': Vg,
                            'Is': Is,
                            'Id': Id,
                            'Ig': Ig
                        })
                        df.to_excel(writer, sheet_name=device_ids[-1], index=False)

                    # 新 device
                    device_id = line.split("[")[1].split("]")[0]
                    device_ids.append(device_id)

                    # 清空缓存
                    Vg, Id, Is, Ig = [], [], [], []

                    # 跳过后面 4 行说明/表头
                    for _ in range(4):
                        f.readline()

                    # 继续读取数据段直到空行
                    while True:
                        pos = f.tell()
                        data_line = f.readline()

                        if data_line == "" or data_line.strip() == "":
                            break

                        cols = data_line.strip().split("\t")
                        if len(cols) < 4:
                            # 兼容空格分隔
                            cols = data_line.strip().split()

                        if len(cols) >= 4:
                            try:
                                Vg.append(float(cols[0]))
                                Id.append(float(cols[1]))
                                Is.append(float(cols[2]))
                                Ig.append(float(cols[3]))
                            except ValueError:
                                pass
                        else:
                            # 非法行，回退并结束当前段
                            f.seek(pos)
                            break

            # 保存最后一个 device
            if len(Vg) > 0 and len(device_ids) > 0:
                df = pd.DataFrame({
                    'Vg': Vg,
                    'Is': Is,
                    'Id': Id,
                    'Ig': Ig
                })
                df.to_excel(writer, sheet_name=device_ids[-1], index=False)

    return device_ids


def parse_it_to_excel(it_path, out_excel):
    """
    解析 sampling/Time 文件 -> It.xlsx
    """
    with open(it_path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()

    # 找到 Time 表头
    header_idx = None
    device_ids = []

    for idx, line in enumerate(lines):
        if line.startswith("Time"):
            header_idx = idx
            for seg in line.split("["):
                if "]" in seg:
                    device_ids.append(seg.split("]")[0])
            break

    if header_idx is None:
        raise ValueError(f"{it_path} 中没有找到 'Time' 表头，可能拿错文件了")

    if len(device_ids) == 0:
        raise ValueError(f"{it_path} 中没有解析出 device ID")

    dn = len(device_ids)
    device_data = {dev: {'Time': [], 'Is': [], 'Id': [], 'Ig': []} for dev in device_ids}

    # header下一行开始是数据
    for line in lines[header_idx + 1:]:
        if line.strip() == "":
            continue

        cols = line.strip().split("\t")
        if len(cols) < dn * 5:
            cols = line.strip().split()

        if len(cols) < dn * 5:
            continue

        for t in range(dn):
            dev = device_ids[t]
            try:
                device_data[dev]['Time'].append(float(cols[t * 5]))
                device_data[dev]['Is'].append(float(cols[t * 5 + 1]))
                device_data[dev]['Id'].append(float(cols[t * 5 + 2]))
                device_data[dev]['Ig'].append(float(cols[t * 5 + 4]))
            except ValueError:
                continue

    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        for dev in device_ids:
            df = pd.DataFrame(device_data[dev])
            df.to_excel(writer, sheet_name=dev, index=False)

    return device_ids


def generate_vt_excel(iv_excel, it_excel, out_excel, device_ids):
    """
    通过 IV 曲线拟合，把 It 中的 Id 映射成 Vt
    """
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        for dev in device_ids:
            IVdata = pd.read_excel(iv_excel, sheet_name=dev, usecols=[0, 2])   # Vg, Id
            Itdata = pd.read_excel(it_excel, sheet_name=dev, usecols=[0, 2])   # Time, Id

            Ttt = Itdata['Time'].astype(float).tolist()
            CurrT1 = Itdata['Id'].astype(float).abs().tolist()

            Volt = IVdata['Vg'].astype(float).tolist()
            Curr = IVdata['Id'].astype(float).tolist()

            if len(Volt) < 2 or len(Curr) < 2:
                raise ValueError(f"{dev}: IV 数据点太少，无法拟合")

            tmin = None
            tmax = None
            curr_min = min(CurrT1)
            curr_max = max(CurrT1)

            for t in range(len(Curr) - 1):
                if Curr[t] < curr_min <= Curr[t + 1]:
                    tmin = t
                if Curr[t] <= curr_max < Curr[t + 1]:
                    tmax = t + 1

            if tmin is None:
                raise ValueError(
                    f"{dev}: 没找到 tmin。sampling 的最小 Id={curr_min:.3e} 不在 IV 范围内"
                )
            if tmax is None:
                raise ValueError(
                    f"{dev}: 没找到 tmax。sampling 的最大 Id={curr_max:.3e} 不在 IV 范围内"
                )

            fit_Vg = Volt[max(0, tmin - 1):min(len(Volt), tmax + 2)]
            fit_Id = Curr[max(0, tmin - 1):min(len(Curr), tmax + 2)]

            if len(fit_Id) < 3:
                raise ValueError(f"{dev}: 拟合区间数据点太少")

            # 初值
            a_guess = 0.01
            b_guess = 0.1
            c_guess = -1e-10

            try:
                popt, pcov = curve_fit(
                    lambda l, a, b, c: a * np.log(l - c) + b,
                    fit_Id,
                    fit_Vg,
                    p0=(a_guess, b_guess, c_guess),
                    maxfev=50000
                )
            except Exception as e:
                raise RuntimeError(f"{dev}: curve_fit 失败，原因：{e}")

            a, b, c = popt

            if min(np.array(CurrT1) - c) <= 0:
                raise ValueError(
                    f"{dev}: 对数拟合中 CurrT1 - c <= 0，无法计算 log"
                )

            Vt = a * np.log(np.array(CurrT1) - c) + b

            df = pd.DataFrame({
                'Time': Ttt,
                'Vt': Vt
            })
            df.to_excel(writer, sheet_name=dev, index=False)


def normalize_vt_excel(vt_excel, out_excel, begin_value, device_ids):
    """
    归一化时间与电压
    """
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        for dev in device_ids:
            Vtdata = pd.read_excel(vt_excel, sheet_name=dev, usecols=[0, 1])

            Vtt = Vtdata['Time'].astype(float).tolist()
            Vtv = Vtdata['Vt'].astype(float).tolist()

            cross_idx = None
            for idx in range(1, len(Vtt)):
                if Vtt[idx] >= begin_value and Vtt[idx - 1] < begin_value:
                    cross_idx = idx
                    break

            if cross_idx is None:
                raise ValueError(
                    f"{dev}: 没找到 Begin={begin_value} 对应的归一化起点"
                )

            # 以第一个 >= Begin 的点作为 t0 / Vt0
            t0 = Vtt[cross_idx]
            v0 = Vtv[cross_idx]

            # 只输出 t0 时刻之后（包含 t0）的数据，去掉 t0 之前的数据
            Vtt_after_t0 = np.array(Vtt[cross_idx:])
            Vtv_after_t0 = np.array(Vtv[cross_idx:])

            Vtt1 = (Vtt_after_t0 - t0) / 60.0      # min
            Vtv1 = (Vtv_after_t0 - v0) * 1000.0    # mV

            df = pd.DataFrame({
                'T-T0': Vtt1,
                'Vt-Vt0': Vtv1
            })
            df.to_excel(writer, sheet_name=dev, index=False)


# ========================= 主程序 =========================

for foldername in os.listdir(FilePath):
    fpt = os.path.join(FilePath, foldername)

    if not os.path.isdir(fpt):
        continue

    print(f"\nProcessing folder: {fpt}")

    try:
        iv_path, it_path, begin_value = classify_txt_files(fpt)
        print("IV file     :", iv_path)
        print("It file     :", it_path)
        print("Begin value :", begin_value)

        IVxls = os.path.join(fpt, 'IV.xlsx')
        Itxls = os.path.join(fpt, 'It.xlsx')
        Vtxls = os.path.join(fpt, 'Vt.xlsx')
        Vtxls1 = os.path.join(fpt, 'Vt1.xlsx')

        # 先创建空工作簿，避免旧文件残留也可省略
        Workbook().save(IVxls)
        Workbook().save(Itxls)
        Workbook().save(Vtxls)
        Workbook().save(Vtxls1)

        device_ids_iv = parse_iv_to_excel(iv_path, IVxls)
        device_ids_it = parse_it_to_excel(it_path, Itxls)

        # 用共同 device
        device_ids = [d for d in device_ids_it if d in device_ids_iv]
        if len(device_ids) == 0:
            raise ValueError(f"{fpt}: IV 和 It 中没有匹配的 device sheet")

        generate_vt_excel(IVxls, Itxls, Vtxls, device_ids)
        normalize_vt_excel(Vtxls, Vtxls1, begin_value, device_ids)

        Vtfile.append(Vtxls1)

    except Exception as e:
        print(f"Error in folder [{fpt}]: {e}")


# ========================= 画图 =========================

plt.figure()

for idx, vt_file in enumerate(Vtfile):
    wb = load_workbook(vt_file, read_only=True, keep_links=False)
    plot_color = color[idx % len(color)]

    for sheet in wb.sheetnames:
        Vtdata = pd.read_excel(vt_file, sheet_name=sheet, usecols=[0, 1])
        Vtt = Vtdata['T-T0'].astype(float).tolist()
        Vtv = Vtdata['Vt-Vt0'].astype(float).tolist()

        label_name = os.path.basename(os.path.dirname(vt_file)) + "_" + sheet
        plt.plot(Vtt, Vtv, color=plot_color, label=label_name)

plt.legend()
plt.xlim(0, 90)
plt.ylim(-10, 150)
plt.xlabel("Time relative to T0 (min)")
plt.ylabel("Vt - Vt0 (mV)")
plt.show()